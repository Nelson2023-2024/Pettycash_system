from openpyxl.styles import Font, PatternFill, Alignment

from services.services import NotificationService, PettyCashAccountService, UserService
from utils.response_provider import ResponseProvider
from utils.common import get_clean_request_data
from django.core.exceptions import ValidationError
from audit.models import Notifications, TransactionLogBase
from datetime import timedelta
from django.utils import timezone
import openpyxl
from django.http import HttpResponse
from audit.models import EventTypes



class PettyCashService:

    @classmethod
    def create_petty_cash_account(cls, request) -> ResponseProvider:
        data = get_clean_request_data(
            request,
            required_fields={
                "name",
                "description",
                "mpesa_phone_number",
                "minimum_threshold",
            },
        )

        name = data.get("name")
        description = data.get("description")
        mpesa_phone_number = data.get("mpesa_phone_number")
        minimum_threshold = data.get("minimum_threshold")

        petty_cash, log = PettyCashAccountService().create_account(
            name,
            description,
            mpesa_phone_number,
            minimum_threshold,
            triggered_by=request.user,
            request=request,
        )

        NotificationService.notify_many(
            transaction_log=log,
            recipients=UserService().get_active_admin_fo_cfo(),
            channel=Notifications.Channel.IN_APP,
        )

        return ResponseProvider.created(
            message=f"{petty_cash.name} account created successfully",
            data=cls._serialize(petty_cash),
        )

    @classmethod
    def get_petty_cash_account(cls, account_id: str):
        petty_cash = PettyCashAccountService().get_by_id(account_id)
        return ResponseProvider.success(data=cls._serialize(petty_cash))

    @classmethod
    def get_all_petty_cash_accounts(cls):
        accounts = PettyCashAccountService().get_active_accounts()
        data = []
        for account in accounts:
            data.append(cls._serialize(account))
        return ResponseProvider.success(data=data)

    @classmethod
    def update_petty_cash_account(cls, request, account_id: str):
        data = get_clean_request_data(
            request,
            allowed_fields={
                "name",
                "description",
                "mpesa_phone_number",
                "minimum_threshold",
                "account_type",
            },
        )

        petty_cash, log = PettyCashAccountService().update_account(
            account_id, data, triggered_by=request.user, request=request
        )
        NotificationService.notify(
            transaction_log=log,
            recipient=request.user,
            channel=Notifications.Channel.IN_APP,
        )
        return ResponseProvider.success(
            message=f"{petty_cash.name} updated successfully",
            data=cls._serialize(petty_cash),
        )

    @staticmethod
    def deactivate_petty_cash_account(request, account_id: str):
        petty_cash, log = PettyCashAccountService().deactivate_account(
            account_id, triggered_by=request.user, request=request
        )

        NotificationService.notify(
            transaction_log=log,
            recipient=request.user,
            channel=Notifications.Channel.IN_APP,
        )
        return ResponseProvider.success(
            message=f"{petty_cash.name} deactivated successfully"
        )

    @classmethod
    def get_account_activity(cls, request) -> ResponseProvider:
        """
        Retrieves all transaction logs related to the active petty cash account.
        Returns a full audit trail including:
        - Who triggered the transaction
        - Money before and after
        - Amount deducted/added
        - Date and time
        - Event type (disbursement, topup, deduction, etc.)
        """
        try:
            account = PettyCashAccountService().get_active_accounts().first()
            if not account:
                return ResponseProvider.not_found(
                    message="No active petty cash account found."
                )

            logs = (
                TransactionLogBase.objects.filter(
                    event_type__code__in=[
                        "petty_cash_balance_deducted",  # money out — expense disbursed
                        "topup_disbursed",  # # money in — topup credited
                    ],
                    metadata__account_id=str(account.id),
                )
                .select_related("triggered_by")
                .order_by("-created_at")
            )

            return ResponseProvider.success(
                data={
                    "account": cls._serialize(account),
                    "activity": [cls._serialize_log(log) for log in logs],
                }
            )
        except Exception as ex:
            return ResponseProvider.handle_exception(ex)

    @classmethod
    def export_account_activity(cls, request) -> HttpResponse:
        """
        Exports petty cash account activity as Excel.

        ?period=weekly  → last 7 days
        ?period=monthly → last 30 days (default)

        Columns: Date, Item, Triggered By, Amount, Transaction Cost, Total Spent, Balance
        """
        try:
            account = PettyCashAccountService().get_active_accounts().first()
            if not account:
                return ResponseProvider.not_found(
                    message="No active petty cash account found."
                )

            period = request.GET.get("period", "monthly")
            days = 7 if period == "weekly" else 30
            date_from = timezone.now() - timedelta(days=days)

            logs = (
                TransactionLogBase.objects.filter(
                    event_type__code__in=[
                        "petty_cash_balance_deducted",
                        "topup_disbursed",
                    ],
                    metadata__account_id=str(account.id),
                    created_at__gte=date_from,
                )
                .select_related("triggered_by", "event_type")
                .order_by("-created_at")
            )

            # ── Build workbook ────────────────────────────────
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{period.capitalize()} Activity"

            # ── Header row ────────────────────────────────────
            headers = [
                "Date", "Item", "Triggered By",
                "Amount (KES)", "Transaction Cost (KES)",
                "Total Spent (KES)", "Balance (KES)",
            ]

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="1a1a2e",
                end_color="1a1a2e",
                fill_type="solid"
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # ── Data rows ─────────────────────────────────────
            for row_num, log in enumerate(logs, 2):
                serialized = cls._serialize_log(log)

                ws.cell(
                    row=row_num,
                    column=1,
                    value=log.created_at.strftime("%d %b %Y %H:%M"),
                )
                ws.cell(row=row_num, column=2, value=serialized.get("item", ""))

                ws.cell(
                    row=row_num,
                    column=3,
                    value=(
                        serialized.get("triggered_by", {}).get("email", "")
                        if serialized.get("triggered_by")
                        else "System"
                    ),
                )

                ws.cell(row=row_num, column=4, value=serialized.get("amount", ""))
                ws.cell(row=row_num, column=5, value=serialized.get("transaction_cost", ""))
                ws.cell(row=row_num, column=6, value=serialized.get("total_spent", ""))
                ws.cell(row=row_num, column=7, value=serialized.get("balance", ""))

            # ── Auto column width ─────────────────────────────
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4

                # ── Log the export as a transaction ───────────────
            # TransactionLogBase.objects.create(
            #         triggered_by=request.user,
            #         event_type=EventTypes.objects.get(code="petty_cash_export"),
            #         metadata={
            #             "account_id": str(account.id),
            #             "period": period,
            #             "export_count": logs.count(),
            #         },
            #         event_message=f"{request.user.email} exported petty cash activity ({period})",
            #     )

            # ── Return as download ────────────────────────────
            filename = f"petty_cash_{period}_{timezone.now().strftime('%Y%m%d')}.xlsx"

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            wb.save(response)
            return response

        except ValidationError as ve:
            return ResponseProvider.bad_request(message=str(ve))

        except Exception as ex:
            return ResponseProvider.handle_exception(ex)

    @staticmethod
    def _serialize(petty_cash) -> dict:
        """
        Converting a Django model → JSON-safe dictionary
        """
        return {
            "id": str(petty_cash.id),
            "name": petty_cash.name,
            "description": petty_cash.description,
            "mpesa_phone_number": petty_cash.mpesa_phone_number,
            "account_type": petty_cash.account_type,
            "current_balance": str(petty_cash.current_balance),
            "minimum_threshold": str(petty_cash.minimum_threshold),
            "is_active": petty_cash.is_active,
            "created_at": str(petty_cash.created_at),
            "updated_at": str(petty_cash.updated_at),
        }

    @staticmethod
    def _serialize_log(log) -> dict:
        metadata = log.metadata or {}

        expense_amount = metadata.get("expense_amount")
        transaction_cost = metadata.get("transaction_cost")
        amount_deducted = metadata.get("amount_deducted")
        topup_amount = metadata.get("amount")

        try:
            total_spent = (
                str(float(expense_amount) + float(transaction_cost))
                if expense_amount and transaction_cost
                else amount_deducted or topup_amount
            )
        except (TypeError, ValueError):
            total_spent = amount_deducted or topup_amount

        # ── item: use expense title if available, else log message ──
        title = metadata.get("title")
        employee_email = metadata.get("employee_email")
        item = (
            f"{title} — {employee_email}"
            if title and employee_email
            else title or log.event_message
        )

        return {
            "date": log.created_at.isoformat(),
            "triggered_by": (
                {
                    "id": str(log.triggered_by.id),
                    "name": f"{log.triggered_by.first_name} {log.triggered_by.last_name}".strip(),
                    "email": log.triggered_by.email,
                }
                if log.triggered_by
                else None
            ),
            "item": item,
            "amount": expense_amount or topup_amount,
            "transaction_cost": transaction_cost or "0",
            "total_spent": total_spent,
            "balance": metadata.get("new_balance"),
            "event_code": log.event_type.code if log.event_type else None,
            "event_name": log.event_type.name if log.event_type else None,
        }
